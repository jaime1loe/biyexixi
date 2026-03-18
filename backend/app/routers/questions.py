from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import desc
from typing import List, Optional
import csv
from io import StringIO
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from urllib.parse import quote

from app.database import get_db
from app.models import Question
from app.schemas import QuestionCreate, QuestionResponse
from app.dependencies import get_current_user

router = APIRouter()


@router.post("/", response_model=QuestionResponse, summary="提问")
async def create_question(
    question: QuestionCreate,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """用户提问接口"""
    # 检查是否已有相同问题
    existing_question = db.query(Question).filter(
        Question.question == question.question
    ).first()

    if existing_question:
        # 如果问题已存在，增加提问次数
        existing_question.ask_count += 1
        db.commit()
        db.refresh(existing_question)
        return existing_question
    else:
        # 创建新问题记录
        db_question = Question(
            user_id=current_user.id,
            question=question.question,
            category=question.category,
            ask_count=1  # 新问题默认提问次数为1
        )
        db.add(db_question)
        db.commit()
        db.refresh(db_question)

        # 这里可以调用AI服务生成答案
        # 暂时返回占位答案
        db_question.answer = "这是一个示例答案。AI功能将在后续阶段实现。"
        db.commit()
        db.refresh(db_question)

        return db_question


@router.get("/my", response_model=List[QuestionResponse], summary="获取我的问题列表")
async def get_my_questions(
    skip: int = 0,
    limit: int = 20,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """获取当前用户的问题列表"""
    questions = db.query(Question).filter(
        Question.user_id == current_user.id
    ).order_by(desc(Question.created_at)).offset(skip).limit(limit).all()
    return questions


@router.get("/export/csv", summary="导出历史问题为CSV")
async def export_questions_csv(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """导出当前用户的历史问题为CSV文件"""
    questions = db.query(Question).filter(
        Question.user_id == current_user.id
    ).order_by(desc(Question.created_at)).all()

    # 创建CSV内容 - 使用 StringIO 然后转换为 BytesIO 并添加 BOM
    import codecs
    text_output = StringIO()
    writer = csv.writer(text_output)
    writer.writerow(["ID", "问题", "答案", "分类", "浏览次数", "创建时间"])

    for q in questions:
        writer.writerow([
            q.id,
            q.question,
            q.answer or "",
            q.category or "",
            q.views,
            q.created_at.strftime("%Y-%m-%d %H:%M:%S")
        ])

    # 获取文本内容并转换为带 BOM 的 BytesIO
    csv_content = text_output.getvalue()
    output = BytesIO()
    output.write(codecs.BOM_UTF8)
    output.write(csv_content.encode('utf-8'))
    output.seek(0)

    # 返回CSV文件
    filename = f"questions_{current_user.username}_{current_user.id}.csv"
    response = StreamingResponse(
        output,
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"
        }
    )
    return response


@router.get("/export/excel", summary="导出历史问题为Excel")
async def export_questions_excel(
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """导出当前用户的历史问题为Excel文件"""
    questions = db.query(Question).filter(
        Question.user_id == current_user.id
    ).order_by(desc(Question.created_at)).all()

    # 创建Excel工作簿
    wb = Workbook()
    ws = wb.active
    ws.title = "问答历史"

    # 定义样式
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    cell_font = Font(size=11)
    alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 设置列宽
    ws.column_dimensions['A'].width = 10   # ID
    ws.column_dimensions['B'].width = 50   # 问题
    ws.column_dimensions['C'].width = 50   # 答案
    ws.column_dimensions['D'].width = 15   # 分类
    ws.column_dimensions['E'].width = 12   # 浏览次数
    ws.column_dimensions['F'].width = 20   # 创建时间

    # 写入表头
    headers = ["序号", "问题", "答案", "分类", "浏览次数", "创建时间"]
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # 写入数据
    for row_num, question in enumerate(questions, 2):
        ws.cell(row=row_num, column=1, value=row_num - 1).border = border
        ws.cell(row=row_num, column=2, value=question.question).alignment = alignment
        ws.cell(row=row_num, column=2).font = cell_font
        ws.cell(row=row_num, column=2).border = border

        ws.cell(row=row_num, column=3, value=question.answer or "").alignment = alignment
        ws.cell(row=row_num, column=3).font = cell_font
        ws.cell(row=row_num, column=3).border = border

        ws.cell(row=row_num, column=4, value=question.category or "").border = border
        ws.cell(row=row_num, column=5, value=question.views).border = border
        ws.cell(row=row_num, column=6, value=question.created_at.strftime("%Y-%m-%d %H:%M:%S")).border = border

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 保存到内存
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # 返回Excel文件
    filename = f"问答历史_{current_user.username}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    # 使用 ASCII 编码的文件名避免 HTTP header 编码问题
    ascii_filename = f"qa_history_{current_user.username}_{datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    response = StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={ascii_filename}"
        }
    )
    return response


@router.get("/categories", summary="获取问题分类列表")
async def get_question_categories(db: Session = Depends(get_db)):
    """获取所有问题分类"""
    categories = db.query(Question.category).distinct().all()
    return {"categories": [c[0] for c in categories if c[0]]}


@router.get("/", response_model=List[QuestionResponse], summary="获取问题列表")
async def get_questions(
    skip: int = 0,
    limit: int = 10,
    category: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """获取问题列表（公开）"""
    query = db.query(Question)
    if category:
        query = query.filter(Question.category == category)
    questions = query.order_by(desc(Question.created_at)).offset(skip).limit(limit).all()
    return questions


@router.get("/{question_id}", response_model=QuestionResponse, summary="获取问题详情")
async def get_question(question_id: int, db: Session = Depends(get_db)):
    """获取问题详情"""
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题不存在"
        )
    # 增加浏览次数
    question.views += 1
    db.commit()
    return question


@router.put("/{question_id}", response_model=QuestionResponse, summary="更新问题分类")
async def update_question(
    question_id: int,
    category: str,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """更新问题分类（用户和管理员）"""
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题不存在"
        )

    # 只有问题所有者或管理员可以修改
    if question.user_id != current_user.id and current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权修改此问题"
        )

    question.category = category
    db.commit()
    db.refresh(question)
    return question


@router.delete("/{question_id}", summary="删除问题")
async def delete_question(
    question_id: int,
    db: Session = Depends(get_db),
    current_user = Depends(get_current_user)
):
    """删除问题"""
    question = db.query(Question).filter(Question.id == question_id).first()
    if not question:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail="问题不存在"
        )

    # 只有问题所有者或管理员可以删除
    if question.user_id != current_user.id and current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN,
            detail="无权删除此问题"
        )

    db.delete(question)
    db.commit()
    return {"message": "删除成功"}
